import json
import logging
import inspect
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from exporters.excel_exporter import ExcelExporter, JOB_FIELDS
from utils.run_paths import (
    build_final_run_name,
    create_run_directory,
    find_run_directory,
    finalize_run_directory,
    flatten_task_artifacts,
    internalize_run_directory,
    keyword_summary,
    latest_run_directory,
    nest_platform_artifacts,
    rewrite_artifact_paths,
    run_file_path,
    sanitize_run_component,
    write_run_config,
)
from utils.storage import JsonlStore
import main


def test_running_directory_is_independent_and_latest_tracks_it(tmp_path):
    when = datetime(2026, 7, 13, 18, 30, 45)
    first = create_run_directory(tmp_path, when)
    second = create_run_directory(tmp_path, when)
    assert first.name == ".running_20260713_183045"
    assert second.name == ".running_20260713_183045_2"
    assert (second / "screenshots").is_dir()
    assert (second / "html").is_dir()
    assert (second / "debug").is_dir()
    assert (tmp_path / "latest").is_symlink()
    assert latest_run_directory(tmp_path) == second.resolve()


def test_one_main_invocation_creates_only_one_run_directory_for_all_keywords():
    source = inspect.getsource(main.main)
    assert source.count("create_run_directory(") == 1


def test_two_keywords_generate_readable_final_name():
    completed = datetime(2026, 7, 13, 19, 10)
    assert build_final_run_name(
        completed, ["交易系统运维", "合规风控"]
    ) == "2026-07-13_19-10_交易系统运维+合规风控"


def test_no_new_jobs_completed_task_has_no_partial_prefix():
    assert build_final_run_name(
        datetime(2026, 7, 13, 19, 10), ["交易系统运维", "合规审核"],
        completed=True, status="completed",
    ) == "2026-07-13_19-10_交易系统运维+合规审核"


def test_more_than_three_keywords_use_other_count():
    assert keyword_summary(["关键词一", "关键词二", "关键词三", "关键词四", "关键词五", "关键词六"]) == (
        "关键词一+关键词二+关键词三+3-more"
    )
    long_name = build_final_run_name(
        datetime(2026, 7, 13, 19, 10), ["甲" * 50, "乙" * 50, "丙" * 50, "丁", "戊"]
    )
    assert len(long_name) <= 100
    assert long_name.endswith("+2-more")


def test_illegal_filename_characters_are_replaced_and_length_is_limited():
    assert sanitize_run_component('交易/系统\\运维:*?"<>|') == "交易_系统_运维_______"
    name = build_final_run_name(datetime(2026, 7, 13, 19, 10), ["超长" * 100])
    assert len(name) <= 100
    assert not any(character in name for character in '/\\:*?"<>|')


def test_same_final_name_gets_numeric_suffix(tmp_path):
    completed = datetime(2026, 7, 13, 19, 10)
    first = create_run_directory(tmp_path, datetime(2026, 7, 13, 19, 0), direct=True)
    second = create_run_directory(tmp_path, datetime(2026, 7, 13, 19, 1), direct=True)
    write_run_config(first / "run_config.json", {"search_keywords": ["测试关键词"]})
    write_run_config(second / "run_config.json", {"search_keywords": ["测试关键词"]})
    final_first = finalize_run_directory(
        first, tmp_path, ["测试关键词"], completed, completed=True
    )
    final_second = finalize_run_directory(
        second, tmp_path, ["测试关键词"], completed, completed=True
    )
    assert final_first.name == "2026-07-13_19-10_测试关键词"
    assert final_second.name == "2026-07-13_19-10_测试关键词_2"


def test_success_and_interrupted_runs_are_renamed(tmp_path):
    completed = datetime(2026, 7, 13, 19, 10)
    success = create_run_directory(tmp_path, datetime(2026, 7, 13, 19, 0), direct=True)
    stopped = create_run_directory(tmp_path, datetime(2026, 7, 13, 19, 1), direct=True)
    write_run_config(success / "run_config.json", {"search_keywords": ["交易系统运维"]})
    write_run_config(stopped / "run_config.json", {"search_keywords": ["合规风控"]})
    final_success = finalize_run_directory(
        success, tmp_path, ["交易系统运维"], completed, completed=True
    )
    final_stopped = finalize_run_directory(
        stopped, tmp_path, ["合规风控"], completed, completed=False
    )
    assert not success.exists()
    assert final_success.name == "2026-07-13_19-10_交易系统运维"
    assert final_stopped.name == "2026-07-13_19-10_incomplete_合规风控"
    assert latest_run_directory(tmp_path) == final_stopped.resolve()


def test_run_id_finds_same_task_before_and_after_rename(tmp_path):
    running = create_run_directory(
        tmp_path, datetime(2026, 7, 13, 19, 0), direct=True
    )
    run_id = running.name.removeprefix(".running_")
    write_run_config(running / "run_config.json", {
        "run_id": run_id,
        "search_keywords": ["交易系统运维"],
    })
    assert find_run_directory(tmp_path, run_id) == running.resolve()
    final_dir = finalize_run_directory(
        running, tmp_path, ["交易系统运维"], datetime(2026, 7, 13, 19, 10),
        completed=True,
    )
    assert find_run_directory(tmp_path, run_id) == final_dir.resolve()


def test_renamed_artifacts_keep_jsonl_paths_and_excel_links_valid(tmp_path):
    logger = logging.getLogger("run-path-test")
    running = create_run_directory(
        tmp_path, datetime(2026, 7, 13, 19, 0), direct=True
    )
    write_run_config(running / "run_config.json", {"search_keywords": ["交易系统运维"]})
    screenshot = running / "screenshots" / "交易系统运维" / "1.png"
    html = running / "html" / "交易系统运维" / "1.html"
    screenshot.parent.mkdir(parents=True)
    html.parent.mkdir(parents=True)
    screenshot.write_bytes(b"png")
    html.write_text("<html></html>", encoding="utf-8")
    records = [{
        "search_keyword": "交易系统运维",
        "search_rank": 1,
        "title": "测试岗位",
        "company": "测试公司",
        "url": "https://www.zhipin.com/job_detail/abc.html",
        "screenshot_path": str(screenshot),
        "html_path": str(html),
    }]
    final_dir = finalize_run_directory(
        running, tmp_path, ["交易系统运维"], datetime(2026, 7, 13, 19, 10),
        completed=True,
    )
    rewrite_artifact_paths(records, running, final_dir, tmp_path)
    store = JsonlStore(final_dir / "jobs.jsonl", logger)
    store.write_all(records)
    ExcelExporter(final_dir / "jobs.xlsx", logger, project_root=final_dir).export(
        records, [], {"status": "completed"}, []
    )

    saved = json.loads((final_dir / "jobs.jsonl").read_text(encoding="utf-8"))
    assert saved["screenshot_path"] == "screenshots/交易系统运维/1.png"
    assert saved["html_path"] == "html/交易系统运维/1.html"
    assert (final_dir / saved["screenshot_path"]).exists()
    assert (final_dir / saved["html_path"]).exists()
    workbook = load_workbook(final_dir / "jobs.xlsx")
    link = workbook["Jobs_Raw"].cell(2, JOB_FIELDS.index("screenshot_path") + 1).hyperlink
    assert link is not None
    assert link.target == str((final_dir / saved["screenshot_path"]).resolve())


def test_platform_artifacts_are_nested_without_changing_boss_capture_code(tmp_path):
    run_dir = create_run_directory(tmp_path, direct=True)
    screenshot = run_dir / "screenshots" / "交易系统运维" / "1.png"
    html = run_dir / "html" / "交易系统运维" / "1.html"
    screenshot.parent.mkdir(parents=True)
    html.parent.mkdir(parents=True)
    screenshot.write_bytes(b"png")
    html.write_text("<html></html>", encoding="utf-8")
    records = [{"screenshot_path": str(screenshot), "html_path": str(html)}]
    nest_platform_artifacts(run_dir, "boss", records)
    assert (run_dir / "screenshots" / "boss" / "交易系统运维" / "1.png").exists()
    assert (run_dir / "html" / "boss" / "交易系统运维" / "1.html").exists()
    assert "/screenshots/boss/" in records[0]["screenshot_path"]


def test_final_task_artifacts_are_flat_and_technical_files_are_internal(tmp_path):
    run_dir = create_run_directory(tmp_path, direct=True)
    screenshot = run_dir / "screenshots" / "liepin" / "交易系统运维" / "old.png"
    html = run_dir / "html" / "liepin" / "交易系统运维" / "old.html"
    screenshot.parent.mkdir(parents=True)
    html.parent.mkdir(parents=True)
    screenshot.write_bytes(b"png")
    html.write_text("<html></html>", encoding="utf-8")
    for filename in (
        "jobs.jsonl", "invalid_records.jsonl", "run_config.json", "task_state.json", "app.log",
    ):
        (run_dir / filename).write_text("{}\n", encoding="utf-8")
    record = {
        "platform": "liepin", "search_keyword": "交易/系统运维", "search_rank": 1,
        "company": "测试*公司", "title": "交易:助理", "job_id": "abcdefgh123456",
        "screenshot_path": str(screenshot), "html_path": str(html),
    }
    flatten_task_artifacts(run_dir, [record])
    internalize_run_directory(run_dir)
    assert Path(record["screenshot_path"]).parent == run_dir / "screenshots"
    assert Path(record["html_path"]).parent == run_dir / "internal" / "html"
    assert Path(record["screenshot_path"]).name.startswith("0001_Liepin*交易_系统运维*")
    assert "abcdefgh" in Path(record["screenshot_path"]).stem
    assert not (run_dir / "screenshots" / "liepin").exists()
    assert run_file_path(run_dir, "jobs.jsonl") == run_dir / "internal" / "jobs.jsonl"
    assert (run_dir / "internal" / "debug").is_dir()


def test_flat_paths_survive_final_rename_and_excel_links(tmp_path):
    logger = logging.getLogger("flat-run-path-test")
    running = create_run_directory(
        tmp_path, datetime(2026, 7, 14, 22, 0), direct=True
    )
    write_run_config(running / "run_config.json", {
        "run_id": running.name.removeprefix(".running_"),
        "search_keywords": ["交易系统运维"],
    })
    screenshot = running / "screenshots" / "liepin" / "交易系统运维" / "old.png"
    html = running / "html" / "liepin" / "交易系统运维" / "old.html"
    screenshot.parent.mkdir(parents=True)
    html.parent.mkdir(parents=True)
    screenshot.write_bytes(b"png")
    html.write_text("<html></html>", encoding="utf-8")
    (running / "jobs.jsonl").touch()
    (running / "app.log").touch()
    record = {
        "platform": "liepin", "search_keyword": "交易系统运维", "search_rank": 1,
        "company": "测试公司", "title": "交易助理", "job_id": "1234567890",
        "url": "https://www.liepin.com/job/1234567890.shtml",
        "screenshot_path": str(screenshot), "html_path": str(html),
    }
    flatten_task_artifacts(running, [record])
    final_dir = finalize_run_directory(
        running, tmp_path, ["交易系统运维"], datetime(2026, 7, 14, 22, 10),
        completed=True,
    )
    rewrite_artifact_paths([record], running, final_dir, tmp_path)
    internal = internalize_run_directory(final_dir)
    JsonlStore(internal / "jobs.jsonl", logger).write_all([record])
    ExcelExporter(final_dir / "jobs.xlsx", logger, project_root=final_dir).export(
        [record], [], {"status": "completed"}, []
    )
    saved = json.loads((internal / "jobs.jsonl").read_text(encoding="utf-8"))
    assert saved["screenshot_path"].startswith("screenshots/")
    assert saved["html_path"].startswith("internal/html/")
    assert (final_dir / saved["screenshot_path"]).exists()
    assert (final_dir / saved["html_path"]).exists()
    link = load_workbook(final_dir / "jobs.xlsx")["Jobs_Raw"].cell(
        2, JOB_FIELDS.index("screenshot_path") + 1
    ).hyperlink
    assert link is not None
    assert link.target == str((final_dir / saved["screenshot_path"]).resolve())
