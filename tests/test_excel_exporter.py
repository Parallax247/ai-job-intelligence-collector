import logging

import pandas as pd
from openpyxl import load_workbook

from exporters.excel_exporter import (
    ExcelExporter, INTERVIEW_FIELDS, JOB_DETAIL_FIELDS, JOB_FIELDS,
)


def test_excel_contains_required_sheets_and_formatting(tmp_path):
    path = tmp_path / "jobs.xlsx"
    ExcelExporter(path, logging.getLogger("test")).export(
        [{"search_keyword": "Python", "search_rank": 1, "title": "工程师", "company": "测试公司",
          "url": "https://www.zhipin.com/job_detail/abc.html",
          "screenshot_path": "data/screenshots/a.png", "benefits": ["五险一金"],
          "screenshot_status": "failed", "screenshot_error": "截图超时",
          "responsibilities": "负责完整系统测试", "requirements": "三年以上经验",
          "jd_text": "完整描述" * 100}],
        [{"search_keyword": "Python", "target_count": 2, "captured_count": 1,
          "processed_count": 1, "valid_count": 1, "invalid_count": 0,
          "duplicate_count": 1, "historical_skipped_count": 0,
          "failed_count": 0, "screenshot_failed_count": 0, "status": "completed",
          "error_message": "", "started_at": "a", "finished_at": "b"}],
        {"status": "completed"},
        [{"title": "", "invalid_reason": "title为空"}],
    )
    workbook = pd.ExcelFile(path)
    assert workbook.sheet_names == [
        "Jobs_Raw", "Job_Details", "Invalid_Records", "Keyword_Summary", "Run_Log",
        "Interview_Feedback"
    ]
    book = load_workbook(path)
    jobs = book["Jobs_Raw"]
    assert [cell.value for cell in jobs[1]] == JOB_FIELDS
    assert "platform" in JOB_FIELDS
    assert jobs.freeze_panes == "A2"
    assert jobs.auto_filter.ref
    assert jobs.cell(2, JOB_FIELDS.index("url") + 1).hyperlink is not None
    assert jobs.cell(2, JOB_FIELDS.index("screenshot_path") + 1).hyperlink is not None
    assert jobs.cell(2, JOB_FIELDS.index("url") + 1).value == "Open job"
    assert jobs.cell(2, JOB_FIELDS.index("screenshot_path") + 1).value == "View screenshot"
    assert jobs.cell(2, JOB_FIELDS.index("title") + 1).alignment.horizontal == "center"
    assert jobs.cell(2, JOB_FIELDS.index("title") + 1).alignment.vertical == "center"
    assert jobs.cell(2, JOB_FIELDS.index("screenshot_status") + 1).alignment.horizontal == "center"
    assert jobs.cell(2, JOB_FIELDS.index("screenshot_status") + 1).value == "failed"
    assert jobs.row_dimensions[2].height == 24
    assert not {"benefits", "responsibilities", "requirements", "jd_text"} & set(JOB_FIELDS)

    details = book["Job_Details"]
    assert [cell.value for cell in details[1]] == JOB_DETAIL_FIELDS
    jd = details.cell(2, JOB_DETAIL_FIELDS.index("jd_text") + 1)
    assert jd.value == "完整描述" * 100
    assert jd.alignment.horizontal == "left"
    assert jd.alignment.vertical == "top"
    assert jd.alignment.wrap_text is True
    assert details.cell(2, JOB_DETAIL_FIELDS.index("title") + 1).alignment.horizontal == "center"
    assert details.row_dimensions[2].height <= 120
    assert [cell.value for cell in book["Interview_Feedback"][1]] == INTERVIEW_FIELDS
    assert "city" in [cell.value for cell in book["Keyword_Summary"][1]]
    summary_headers = [cell.value for cell in book["Keyword_Summary"][1]]
    assert {"processed_count", "valid_count", "invalid_count", "status", "error_message"} <= set(
        summary_headers
    )
    status_column = summary_headers.index("status") + 1
    assert book["Keyword_Summary"].cell(2, status_column).value == "completed"
