from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# Jobs_Raw contains compact fields for filtering; long text is stored in Job_Details.
JOB_FIELDS = [
    "platform", "search_keyword", "search_rank", "title", "company", "salary", "city", "district",
    "experience", "education", "recruiter", "company_size", "company_industry",
    "financing_stage", "url", "screenshot_path", "screenshot_status", "screenshot_error",
    "captured_at",
]

JOB_DETAIL_FIELDS = [
    "search_keyword", "search_rank", "title", "company", "benefits", "responsibilities",
    "requirements", "jd_text", "url",
]

ALL_JOB_FIELDS = [
    "platform", "search_keyword", "search_rank", "title", "company", "salary", "city", "district",
    "experience", "education", "benefits", "responsibilities", "requirements", "jd_text",
    "recruiter", "company_size", "company_industry", "financing_stage", "url",
    "screenshot_path", "screenshot_status", "screenshot_error", "captured_at",
]

INVALID_FIELDS = list(dict.fromkeys([
    *ALL_JOB_FIELDS, "job_id", "html_path", "invalid_reason", "invalid_source", "invalidated_at",
]))

INTERVIEW_FIELDS = [
    "company", "title", "apply_date", "resume_result", "interview_date", "interview_round",
    "actual_questions", "weak_points", "result", "offered_salary", "personal_rating", "notes",
]


class ExcelExporter:
    def __init__(self, path: Path, logger, project_root: Path | None = None):
        self.path, self.logger = path, logger
        self.project_root = Path(project_root) if project_root is not None else path.parent.parent

    def export(self, jobs: list[dict[str, Any]], keyword_summary: list[dict[str, Any]],
               run_log: dict[str, Any], invalid_records: list[dict[str, Any]] | None = None) -> None:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        raw_rows = [self._row(job, JOB_FIELDS) for job in jobs]
        detail_rows = [self._row(job, JOB_DETAIL_FIELDS) for job in jobs]
        invalid_rows = [self._row(job, INVALID_FIELDS) for job in (invalid_records or [])]
        summary_fields = [
            "search_keyword", "target_count", "city", "processed_count", "valid_count",
            "invalid_count", "captured_count", "duplicate_count", "historical_skipped_count",
            "failed_count", "screenshot_failed_count", "status", "error_message",
            "started_at", "finished_at",
        ]

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:
            pd.DataFrame(raw_rows, columns=JOB_FIELDS).to_excel(
                writer, sheet_name="Jobs_Raw", index=False
            )
            pd.DataFrame(detail_rows, columns=JOB_DETAIL_FIELDS).to_excel(
                writer, sheet_name="Job_Details", index=False
            )
            pd.DataFrame(invalid_rows, columns=INVALID_FIELDS).to_excel(
                writer, sheet_name="Invalid_Records", index=False
            )
            pd.DataFrame(keyword_summary, columns=summary_fields).to_excel(
                writer, sheet_name="Keyword_Summary", index=False
            )
            pd.DataFrame([run_log]).to_excel(writer, sheet_name="Run_Log", index=False)
            pd.DataFrame(columns=INTERVIEW_FIELDS).to_excel(
                writer, sheet_name="Interview_Feedback", index=False
            )
            self._format_workbook(writer.book)

        self.logger.info(
            "Excel exported: %s (%d valid, %d invalid)", self.path, len(raw_rows), len(invalid_rows)
        )

    @staticmethod
    def _row(record: dict[str, Any], fields: list[str]) -> dict[str, Any]:
        row = {field: record.get(field, "") for field in fields}
        if isinstance(row.get("benefits"), list):
            row["benefits"] = "、".join(
                str(value) for value in row["benefits"] if str(value).strip()
            )
        return row

    def _format_workbook(self, workbook) -> None:
        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            if worksheet.max_column:
                worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.row_dimensions[1].height = 24
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        self._format_jobs_raw(workbook["Jobs_Raw"])
        self._format_job_details(workbook["Job_Details"])
        self._format_invalid(workbook["Invalid_Records"])

        feedback = workbook["Interview_Feedback"]
        for index, field in enumerate(INTERVIEW_FIELDS, 1):
            feedback.column_dimensions[get_column_letter(index)].width = 24 if field != "notes" else 45

    def _format_jobs_raw(self, worksheet) -> None:
        widths = {
            "platform": 12, "search_keyword": 24, "search_rank": 12, "title": 32, "company": 28,
            "salary": 14, "city": 10, "district": 14, "experience": 14, "education": 12,
            "recruiter": 18, "company_size": 14, "company_industry": 20,
            "financing_stage": 14, "url": 14, "screenshot_path": 14,
            "screenshot_status": 16, "screenshot_error": 45, "captured_at": 24,
        }
        header_map = self._header_map(worksheet)
        for field, width in widths.items():
            worksheet.column_dimensions[get_column_letter(header_map[field])].width = width
        centered_fields = {
            "platform", "search_keyword", "search_rank", "title", "company", "salary", "city", "district",
            "experience", "education", "recruiter", "company_size", "company_industry",
            "financing_stage", "screenshot_status", "captured_at",
        }
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 24
            for field in centered_fields:
                worksheet.cell(row, header_map[field]).alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            worksheet.cell(row, header_map["screenshot_error"]).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )
        self._add_links(worksheet, header_map, friendly_labels=True)

    def _format_job_details(self, worksheet) -> None:
        header_map = self._header_map(worksheet)
        widths = {
            "search_keyword": 24, "search_rank": 12, "title": 32, "company": 28,
            "benefits": 45, "responsibilities": 55, "requirements": 55, "jd_text": 60,
            "url": 18,
        }
        for field, width in widths.items():
            worksheet.column_dimensions[get_column_letter(header_map[field])].width = width
        long_fields = ("benefits", "responsibilities", "requirements", "jd_text")
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 120
            for field in ("title", "company"):
                worksheet.cell(row, header_map[field]).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            for field in long_fields:
                worksheet.cell(row, header_map[field]).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
        self._add_links(worksheet, header_map, friendly_labels=True)

    def _format_invalid(self, worksheet) -> None:
        header_map = self._header_map(worksheet)
        for field in ("benefits", "responsibilities", "requirements", "jd_text", "invalid_reason"):
            column = header_map[field]
            worksheet.column_dimensions[get_column_letter(column)].width = 60
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row, column).alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )
        self._add_links(worksheet, header_map, friendly_labels=False)

    @staticmethod
    def _header_map(worksheet) -> dict[str, int]:
        return {cell.value: cell.column for cell in worksheet[1]}

    def _add_links(self, worksheet, header_map: dict[str, int], friendly_labels: bool) -> None:
        labels = {"url": "Open job", "screenshot_path": "View screenshot"}
        for field in ("url", "screenshot_path"):
            column = header_map.get(field)
            if not column:
                continue
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row, column)
                value = str(cell.value or "").strip()
                if not value:
                    continue
                if field == "url":
                    target = value
                else:
                    path = Path(value)
                    target = str(path if path.is_absolute() else (self.project_root / path).resolve())
                cell.hyperlink = target
                if friendly_labels:
                    cell.value = labels[field]
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="center", vertical="center")
